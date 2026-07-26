# -*- coding: utf-8 -*-
"""互联网医疗中心平台 · 项目状态报告 Excel 生成器"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from datetime import datetime
import os

# ── 常量 ──────────────────────────────────────────────
OUTPUT = os.path.join(os.path.dirname(__file__), "互联网医疗中心平台_项目状态报告.xlsx")
TODAY = "2026-07-26"

# ── 配色方案 ──────────────────────────────────────────
C_DARK   = "1B3A5C"   # 深蓝标题栏
C_MID    = "2E6BA8"   # 中蓝
C_LIGHT  = "D6E4F0"   # 浅蓝表头背景
C_ACCENT = "E8F1FA"   # 极浅蓝交替行
C_WHITE  = "FFFFFF"
C_GREEN  = "27AE60"
C_RED    = "E74C3C"
C_AMBER  = "F39C12"
C_LIGHT_GREEN = "E8F8F0"
C_LIGHT_RED   = "FDEDEC"
C_LIGHT_AMBER = "FEF5E7"
C_GREY_BG = "F5F5F5"

# ── 字体 ──────────────────────────────────────────────
F_TITLE  = Font(name="Arial", size=18, bold=True, color=C_WHITE)
F_SUB    = Font(name="Arial", size=11, color=C_WHITE)
F_H1     = Font(name="Arial", size=14, bold=True, color=C_DARK)
F_H2     = Font(name="Arial", size=12, bold=True, color=C_WHITE)
F_TH     = Font(name="Arial", size=10, bold=True, color=C_DARK)
F_TD     = Font(name="Arial", size=10, color="333333")
F_TD_B   = Font(name="Arial", size=10, bold=True, color="333333")
F_KPI    = Font(name="Arial", size=24, bold=True, color=C_DARK)
F_KPI_L  = Font(name="Arial", size=10, color="666666")
F_NOTE   = Font(name="Arial", size=9, italic=True, color="888888")

# ── 填充 ──────────────────────────────────────────────
FILL_DARK   = PatternFill("solid", fgColor=C_DARK)
FILL_MID    = PatternFill("solid", fgColor=C_MID)
FILL_LIGHT  = PatternFill("solid", fgColor=C_LIGHT)
FILL_ACCENT = PatternFill("solid", fgColor=C_ACCENT)
FILL_GREEN  = PatternFill("solid", fgColor=C_LIGHT_GREEN)
FILL_RED    = PatternFill("solid", fgColor=C_LIGHT_RED)
FILL_AMBER  = PatternFill("solid", fgColor=C_LIGHT_AMBER)
FILL_WHITE  = PatternFill("solid", fgColor=C_WHITE)
FILL_GREY   = PatternFill("solid", fgColor=C_GREY_BG)

# ── 边框 ──────────────────────────────────────────────
thin = Side(style="thin", color="CCCCCC")
B_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── 对齐 ──────────────────────────────────────────────
A_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
A_LEFT_T = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ── 工具函数 ──────────────────────────────────────────
def merge_cell(ws, cell_range, value, font, fill=None, align=A_CENTER, border=None):
    ws.merge_cells(cell_range)
    first = cell_range.split(":")[0]
    c = ws[first]
    c.value = value
    c.font = font
    if fill: c.fill = fill
    c.alignment = align
    if border:
        for row in ws[cell_range]:
            for cell in row:
                cell.border = border

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def apply_style(ws, row, col, font=None, fill=None, align=None, border=None):
    c = ws.cell(row=row, column=col)
    if font: c.font = font
    if fill: c.fill = fill
    if align: c.alignment = align
    if border: c.border = border
    return c

def write_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = F_TH
        c.fill = FILL_LIGHT
        c.alignment = A_CENTER
        c.border = B_ALL
    ws.row_dimensions[row].height = 28

def write_data_row(ws, row, data, start_col=1, alt=False, status_col=None):
    fill = FILL_ACCENT if alt else FILL_WHITE
    for i, val in enumerate(data):
        c = ws.cell(row=row, column=start_col + i, value=val)
        c.font = F_TD
        c.fill = fill
        c.alignment = A_LEFT_T if isinstance(val, str) and len(val) > 15 else A_CENTER
        c.border = B_ALL
    if status_col is not None:
        sc = ws.cell(row=row, column=start_col + status_col)
        sv = str(data[status_col]) if data[status_col] else ""
        if "✅" in sv or "已完成" in sv or "PASS" in sv:
            sc.fill = FILL_GREEN
            sc.font = Font(name="Arial", size=10, bold=True, color="1B7A3D")
        elif "🔴" in sv or "❌" in sv or "Not Done" in sv:
            sc.fill = FILL_RED
            sc.font = Font(name="Arial", size=10, bold=True, color="A93226")
        elif "🟡" in sv or "⚠️" in sv or "Partial" in sv or "部分" in sv:
            sc.fill = FILL_AMBER
            sc.font = Font(name="Arial", size=10, bold=True, color="B7780B")
        elif "🟢" in sv:
            sc.fill = FILL_GREEN
            sc.font = Font(name="Arial", size=10, bold=True, color="1B7A3D")
    ws.row_dimensions[row].height = 30

# ═══════════════════════════════════════════════════════
# 创建工作簿
# ═══════════════════════════════════════════════════════
wb = Workbook()

# ─────────────────────────────────────────────────────
# Sheet 1: 项目概况仪表板
# ─────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "项目概况仪表板"
ws1.sheet_view.showGridLines = False
set_col_widths(ws1, [3, 22, 22, 22, 22, 22, 22, 3])

# 标题栏
for r in range(1, 4):
    for c in range(1, 9):
        ws1.cell(row=r, column=c).fill = FILL_DARK
ws1.row_dimensions[1].height = 8
ws1.row_dimensions[2].height = 36
ws1.row_dimensions[3].height = 22
merge_cell(ws1, "B2:G2", "互联网医疗中心平台 · 项目状态报告", F_TITLE, FILL_DARK, A_CENTER)
merge_cell(ws1, "B3:G3", f"报告日期：{TODAY}　|　文档版本：V1.0　|　基线：PRD V2.0 / 技术架构 V1.2　|　编制：项目管理办公室", F_SUB, FILL_DARK, A_CENTER)

# ── KPI 卡片区 ──
ws1.row_dimensions[5].height = 10

kpi_data = [
    ("整体完成度", "≈70%", "后端数据域", C_AMBER),
    ("上线就绪度", "未就绪", "缺合规资质", C_RED),
    ("机制验证", "21/21", "全部 PASS", C_GREEN),
    ("风险等级", "高", "9项关键缺口", C_RED),
    ("阶段状态", "P3完成", "代码级交付", C_AMBER),
    ("PRD覆盖", "≈33%", "前台缺口大", C_RED),
]

row_start = 6
for idx, (label, value, sub, color) in enumerate(kpi_data):
    col = 2 + (idx % 3) * 2
    row = row_start + (idx // 3) * 4
    
    # KPI 值
    merge_cell(ws1, f"{get_column_letter(col)}{row}:{get_column_letter(col+1)}{row}",
               value, Font(name="Arial", size=22, bold=True, color=color), FILL_WHITE, A_CENTER)
    ws1.row_dimensions[row].height = 40
    # KPI 标签
    merge_cell(ws1, f"{get_column_letter(col)}{row+1}:{get_column_letter(col+1)}{row+1}",
               label, F_KPI_L, FILL_WHITE, A_CENTER)
    ws1.row_dimensions[row+1].height = 18
    # KPI 备注
    merge_cell(ws1, f"{get_column_letter(col)}{row+2}:{get_column_letter(col+1)}{row+2}",
               sub, Font(name="Arial", size=9, color="999999"), FILL_WHITE, A_CENTER)
    ws1.row_dimensions[row+2].height = 16
    
    # 边框
    for r2 in range(row, row+3):
        for c2 in range(col, col+2):
            ws1.cell(row=r2, column=c2).border = Border(
                left=Side(style="medium", color=color),
                right=Side(style="thin", color="DDDDDD"),
                top=Side(style="thin", color="DDDDDD"),
                bottom=Side(style="thin", color="DDDDDD"),
            )

# ── 一句话结论 ──
concl_row = 15
ws1.row_dimensions[concl_row].height = 8
merge_cell(ws1, f"B{concl_row+1}:G{concl_row+1}", "总体结论", F_H2, FILL_MID, A_LEFT)
ws1.row_dimensions[concl_row+1].height = 26
merge_cell(ws1, f"B{concl_row+2}:G{concl_row+3}",
           '代码交付基本完成（P0-P3 平台治理与数据中台后端地基已就位，机制级验证 21/21 PASS），'
           '但「上线就绪度 = 未就绪」：缺等保三级测评、互联网医院/药品经营等执业许可（硬门槛）、'
           '真实第三方对接与运行时验证。当前交付物是「可内部研发验证的后端骨架」，'
           '而非「可对外运营的互联网医疗产品」。',
           Font(name="Arial", size=10, color="333333"), FILL_GREY, A_LEFT_T)
ws1.row_dimensions[concl_row+2].height = 24
ws1.row_dimensions[concl_row+3].height = 24

# ── 里程碑状态 ──
ms_row = concl_row + 5
merge_cell(ws1, f"B{ms_row}:G{ms_row}", "里程碑状态", F_H2, FILL_MID, A_LEFT)
ws1.row_dimensions[ms_row].height = 26

write_header_row(ws1, ms_row+1, ["阶段", "目标", "状态", "关键证据", "进度", "风险"])
ms_items = [
    ("P0 主链路", "backend 核心业务路由真实落地并验证", "✅ 已完成", "路由>50；verify_p3 全绿", "100%", "低"),
    ("P1 前后端联调", "admin-web / mp-taro 真实接口对接", "🟢 部分完成", "admin-web 7路由真实接后端", "60%", "中"),
    ("P2 AI与数据闭环", "RAG可解释/反馈闭环/CH经营宽表", "✅ 已完成", "ai-service RAG + pred_log + Grafana", "100%", "低"),
    ("P3 平台治理", "plat真实CRUD/RLS/种子/合理用药/合规", "✅ 已完成（代码层）", "plat CRUD + store_scope + seed + rx_engine", "95%", "中"),
]
for i, item in enumerate(ms_items):
    r = ms_row + 2 + i
    write_data_row(ws1, r, list(item), alt=(i % 2 == 1), status_col=2)
    # 进度条模拟
    pct_cell = ws1.cell(row=r, column=6)
    pct_val = item[4]
    pct_num = int(pct_val.replace("%", ""))
    pct_cell.value = pct_val
    if pct_num >= 90:
        pct_cell.fill = FILL_GREEN
        pct_cell.font = Font(name="Arial", size=10, bold=True, color="1B7A3D")
    elif pct_num >= 50:
        pct_cell.fill = FILL_AMBER
        pct_cell.font = Font(name="Arial", size=10, bold=True, color="B7780B")
    else:
        pct_cell.fill = FILL_RED
        pct_cell.font = Font(name="Arial", size=10, bold=True, color="A93226")

# ── 各模块成熟度 ──
mod_row = ms_row + 2 + len(ms_items) + 1
merge_cell(ws1, f"B{mod_row}:G{mod_row}", "各模块成熟度评估", F_H2, FILL_MID, A_LEFT)
ws1.row_dimensions[mod_row].height = 26

write_header_row(ws1, mod_row+1, ["模块", "状态", "完成度", "关键说明", "验证状态", "优先级"])
modules = [
    ("backend", "🟢 最成熟", "95%", "50+路由，端到端httpx验证PASS", "已验证", "P1"),
    ("admin-web", "🟢 已接接口", "70%", "7路由真实接后端，tsc 0错误", "部分验证", "P3"),
    ("mp-taro", "🔴 空壳", "10%", "5个占位页，无真实业务流", "未验证", "P3"),
    ("ai-service", "✅ 最成熟", "95%", "3接口真实推理，RAG+反馈闭环", "已验证", "P5"),
    ("infra", "🟡 较完整", "65%", "docker-compose全编排，nginx/SQL占位", "部分验证", "P5"),
]
for i, item in enumerate(modules):
    r = mod_row + 2 + i
    write_data_row(ws1, r, list(item), alt=(i % 2 == 1), status_col=1)

# ── PRD覆盖矩阵摘要 ──
prd_row = mod_row + 2 + len(modules) + 1
merge_cell(ws1, f"B{prd_row}:G{prd_row}", "PRD V2.0 功能覆盖摘要", F_H2, FILL_MID, A_LEFT)
ws1.row_dimensions[prd_row].height = 26

write_header_row(ws1, prd_row+1, ["PRD模块", "覆盖状态", "完成度", "判定", "—", "—"])
prd_items = [
    ("互联网医院主线", "后端部分完成，前端空壳", "20%", "⚠️ Partial", "", ""),
    ("健康数据中台主线", "后端+部分前端完成", "65%", "✅ 后端+部分前端", "", ""),
    ("调理师工作台", "无独立模块", "0%", "❌ Not Done", "", ""),
    ("合规大脑", "仅审计日志+RLS", "25%", "⚠️ Partial", "", ""),
    ("数据资产管理", "仅资产目录", "30%", "❌ Not Done", "", ""),
    ("平台监管看板", "仅audit_logs列表", "15%", "⚠️ Partial", "", ""),
]
for i, item in enumerate(prd_items):
    r = prd_row + 2 + i
    write_data_row(ws1, r, list(item), alt=(i % 2 == 1), status_col=3)
    # 合并最后两列
    ws1.merge_cells(f"F{r}:G{r}")
    ws1.cell(row=r, column=6).value = ""
    ws1.cell(row=r, column=6).fill = FILL_ACCENT if i % 2 == 1 else FILL_WHITE
    ws1.cell(row=r, column=6).border = B_ALL

# ─────────────────────────────────────────────────────
# Sheet 2: 本周完成事项
# ─────────────────────────────────────────────────────
ws2 = wb.create_sheet("本周完成事项")
ws2.sheet_view.showGridLines = False
set_col_widths(ws2, [3, 6, 16, 14, 45, 12, 10, 3])

# 标题
for r in range(1, 4):
    for c in range(1, 9):
        ws2.cell(row=r, column=c).fill = FILL_DARK
ws2.row_dimensions[1].height = 8
ws2.row_dimensions[2].height = 36
ws2.row_dimensions[3].height = 22
merge_cell(ws2, "B2:G2", "本周完成事项", F_TITLE, FILL_DARK, A_CENTER)
merge_cell(ws2, "B3:G3", f"报告周期：2026-07-20 ~ {TODAY}　|　对应阶段：P3 平台治理与生产就绪", F_SUB, FILL_DARK, A_CENTER)

# 表头
write_header_row(ws2, 5, ["#", "模块", "阶段", "完成事项", "验证结果", "负责人"])

completed = [
    (1, "backend", "P3", "P3 平台治理完整交付：plat ai_models/data_assets 由占位升级为真实 CRUD（RBAC + 审计）", "21/21 PASS", "后端组"),
    (2, "backend", "P3", "RLS 行级隔离实现：store_scope + JWT store_id 声明 + dev-token 可签发带门店令牌，8 个 mt 端点注入", "机制级 PASS", "后端组"),
    (3, "backend", "P3", "app/db/seed.py 幂等种子：3 门店 + 调理师 + 预置 3 模型/4 资产", "结构验证", "后端组"),
    (4, "backend", "P3", "合理用药引擎 services/rx_engine：抽象层 + Mock 校验，开方前置校验降级不阻断", "Mock PASS", "后端组"),
    (5, "backend", "P3", "等保三级/执业许可合规清单输出（md + docx）", "已交付", "后端组"),
    (6, "backend", "P3", "鉴权强制 + 审计真实化：全部业务路由注入 current_user/require_role；write_audit 的 actor_id/role 取自 JWT", "9/9 PASS", "后端组"),
    (7, "backend", "安全", "jwt_secret 加固：默认值改 32 字节强随机串，InsecureKeyLengthWarning 已消除", "已修复", "后端组"),
    (8, "backend", "P2", "ClickHouse 经营宽表：services/store_metrics.py（PG 按日预聚合写入 CH，幂等）+ mt/store-metrics 接口 + Grafana 看板", "已交付", "后端组"),
    (9, "backend", "P1", "P1 复购/风险回写：调用 ai-service repurchase-prediction v7 / risk-profile v5 真实推理并落库，AI 不可用优雅降级", "已验证", "后端组"),
    (10, "admin-web", "P3", "admin-web 登录入口：登录页（dev 下 dev-token 按角色签发 JWT）+ AuthRoute 守卫 + 401 清理跳转", "tsc 0错误", "前端组"),
    (11, "admin-web", "P1", "真实接口对接：services/mt.ts + services/request.ts（axios 封装、JWT 注入、401 清理）= 真实接口对接", "tsc 0错误", "前端组"),
    (12, "admin-web", "P1", "真实页面交付：customer（列表+授权）、repurchase（发起+历史）、risk（发起+历史）、data（门店经营看板）、plat（审计+AI模型+数据资产）、login", "已交付", "前端组"),
    (13, "ai-service", "P2", "RAG 可解释性实现：plan-recommend 接入 app/rag/retriever.py（内置疼痛/调理知识语料关键词检索），rationale 引用真实知识条目", "已验证", "AI组"),
    (14, "ai-service", "P2", "AI 反馈闭环实现：业务回写落 plat_model_pred_log(adopted=pending) + plat/model-pred-logs 采纳/驳回接口（RBAC + 审计）；model_id 可空", "已验证", "AI组"),
    (15, "ai-service", "P2", "model_loader.py 修复：torch DLL 崩溃、version int→str 两处 bug 修复", "已修复", "AI组"),
    (16, "infra", "P3", "docker-compose.yml 完整编排：PG(+pgvector) / Redis / Mongo / ClickHouse / MinIO / Jaeger / Prometheus / Grafana", "已交付", "运维组"),
    (17, "infra", "P3", "k8s/helm 基础模板 + pre-commit + README", "已交付", "运维组"),
]
for i, item in enumerate(completed):
    r = 6 + i
    write_data_row(ws2, r, list(item), alt=(i % 2 == 1))
    ws2.cell(row=r, column=1).font = F_TD_B
    ws2.cell(row=r, column=1).alignment = A_CENTER
    ws2.row_dimensions[r].height = 36

# 汇总行
sum_row = 6 + len(completed) + 1
merge_cell(ws2, f"B{sum_row}:D{sum_row}", f"合计：{len(completed)} 项已完成", F_TD_B, FILL_LIGHT, A_CENTER, B_ALL)
merge_cell(ws2, f"E{sum_row}:G{sum_row}", "验证状态：21/21 机制级验证 PASS · tsc --noEmit 0 错误", F_TD_B, FILL_GREEN, A_CENTER, B_ALL)
ws2.row_dimensions[sum_row].height = 28

# ─────────────────────────────────────────────────────
# Sheet 3: 下周计划
# ─────────────────────────────────────────────────────
ws3 = wb.create_sheet("下周计划")
ws3.sheet_view.showGridLines = False
set_col_widths(ws3, [3, 6, 12, 28, 30, 10, 10, 10, 3])

for r in range(1, 4):
    for c in range(1, 10):
        ws3.cell(row=r, column=c).fill = FILL_DARK
ws3.row_dimensions[1].height = 8
ws3.row_dimensions[2].height = 36
ws3.row_dimensions[3].height = 22
merge_cell(ws3, "B2:H2", "下周计划（P3-B / P4 排期）", F_TITLE, FILL_DARK, A_CENTER)
merge_cell(ws3, "B3:H3", "排期原则：阻塞上线 → 业务价值 → 验证　|　按优先级 P0–P5 排序", F_SUB, FILL_DARK, A_CENTER)

write_header_row(ws3, 5, ["优先级", "阶段", "计划项", "关键交付物", "预计工时(人天)", "负责人", "状态"])

plan_items = [
    ("P0", "合规硬门槛", "等保三级定级备案与测评排期", "定级备案材料 + 测评机构对接 + 排期表", "5", "合规组", "未启动"),
    ("P0", "合规硬门槛", "申办医疗机构执业许可证 + 互联网医院准入", "许可证申请材料 + 提交回执", "10", "合规组", "未启动"),
    ("P0", "合规硬门槛", "药品经营许可证 + ICP证 + 医师多点执业备案", "全套资质申请清单与时间表", "8", "合规组", "未启动"),
    ("P1", "运行时验证", "alembic upgrade head + 跑 seed.py", "数据库初始化成功 + 种子数据入库", "2", "后端组", "未启动"),
    ("P1", "运行时验证", "用带 store_id 的 dev-token 真实联调 RLS", "8端点RLS真实隔离验证报告", "3", "后端组", "未启动"),
    ("P1", "运行时验证", "验证 pred_log / CH 写入", "AI推理日志 + ClickHouse宽表落库验证", "2", "AI组", "未启动"),
    ("P2", "第三方对接", "HttpRxEngine 接真实合理用药供应商", "真实处方前置校验可用", "5", "后端组", "未启动"),
    ("P2", "第三方对接", "微信支付 / 短信对接（第14章）", "支付链路 + 验证码短信可用", "8", "后端组", "未启动"),
    ("P2", "第三方对接", "开方 patient 信息真实传入（孕期/过敏/日剂量）", "处方安全合规校验可用", "2", "后端组", "未启动"),
    ("P3", "业务前台", "小程序 2C/医生端真实业务流", "首页/找医生/复诊开药/个人中心完整流程", "15", "前端组", "未启动"),
    ("P3", "业务前台", "调理师工作台（PRD 3.5）", "我的客户/方案执行/治疗录入/依从性", "10", "前端组", "未启动"),
    ("P3", "业务前台", "医院管理后台补齐（PRD 3.3）", "科室/药房/药师审核配置", "8", "前端组", "未启动"),
    ("P3", "业务前台", "平台监管看板（PRD 3.7）", "业务总览/异常预警/产品销售监管", "8", "前端组", "未启动"),
    ("P4", "安全闭环", "双因子认证", "管理员2FA登录", "3", "后端组", "未启动"),
    ("P4", "安全闭环", "落库敏感字段加密（KMS）", "敏感数据加密存储方案落地", "5", "后端组", "未启动"),
    ("P4", "安全闭环", "审计日志防篡改（WORM）", "审计日志不可变存储", "3", "后端组", "未启动"),
    ("P5", "性能质量", "压测（Locust / k6）", "核心接口性能基线报告", "5", "测试组", "未启动"),
    ("P5", "性能质量", "ClickHouse 验证 + AI推理响应时间测试", "OLAP + AI性能验证报告", "3", "测试组", "未启动"),
    ("P5", "性能质量", "WAF / CSRF 部署验证", "安全防护部署验证报告", "3", "运维组", "未启动"),
]

for i, item in enumerate(plan_items):
    r = 6 + i
    write_data_row(ws3, r, list(item), alt=(i % 2 == 1))
    # 优先级颜色
    pri_cell = ws3.cell(row=r, column=1)
    pri_val = item[0]
    if pri_val == "P0":
        pri_cell.fill = FILL_RED
        pri_cell.font = Font(name="Arial", size=10, bold=True, color="A93226")
    elif pri_val == "P1":
        pri_cell.fill = FILL_AMBER
        pri_cell.font = Font(name="Arial", size=10, bold=True, color="B7780B")
    elif pri_val == "P2":
        pri_cell.fill = PatternFill("solid", fgColor="FCF3CF")
        pri_cell.font = Font(name="Arial", size=10, bold=True, color="7D6608")
    else:
        pri_cell.fill = FILL_LIGHT
        pri_cell.font = Font(name="Arial", size=10, bold=True, color=C_DARK)
    pri_cell.alignment = A_CENTER
    ws3.row_dimensions[r].height = 32

# 合计行
sum_row3 = 6 + len(plan_items) + 1
total_days = sum(int(x[4]) for x in plan_items)
merge_cell(ws3, f"B{sum_row3}:D{sum_row3}", f"合计：{len(plan_items)} 项计划", F_TD_B, FILL_LIGHT, A_CENTER, B_ALL)
merge_cell(ws3, f"E{sum_row3}:E{sum_row3}", "", F_TD_B, FILL_LIGHT, A_CENTER, B_ALL)
merge_cell(ws3, f"F{sum_row3}:F{sum_row3}", f"{total_days}", F_TD_B, FILL_LIGHT, A_CENTER, B_ALL)
merge_cell(ws3, f"G{sum_row3}:H{sum_row3}", f"总工时：{total_days} 人天", F_TD_B, FILL_LIGHT, A_CENTER, B_ALL)
ws3.row_dimensions[sum_row3].height = 28

# ─────────────────────────────────────────────────────
# Sheet 4: 风险与阻塞项
# ─────────────────────────────────────────────────────
ws4 = wb.create_sheet("风险与阻塞项")
ws4.sheet_view.showGridLines = False
set_col_widths(ws4, [3, 6, 12, 10, 35, 30, 12, 14, 3])

for r in range(1, 4):
    for c in range(1, 10):
        ws4.cell(row=r, column=c).fill = FILL_DARK
ws4.row_dimensions[1].height = 8
ws4.row_dimensions[2].height = 36
ws4.row_dimensions[3].height = 22
merge_cell(ws4, "B2:H2", "风险与阻塞项登记簿", F_TITLE, FILL_DARK, A_CENTER)
merge_cell(ws4, "B3:H3", "来源：对抗式审查风险缺口（第六章）· 共 9 项关键风险", F_SUB, FILL_DARK, A_CENTER)

write_header_row(ws4, 5, ["#", "类别", "等级", "风险描述", "影响 / 后果", "状态", "缓解措施"])

risks = [
    (1, "数据隔离", "高", "RLS 仅代码级，未运行时验证：store_scope 子查询、各端点实际 SQL 过滤未经真实 PG 执行验证", "门店间数据可能泄漏，合规审计不通过", "待处理", "P1 联调阶段用带 store_id 的 dev-token 真实验证"),
    (2, "数据就绪", "高", "种子数据从未执行：seed.py 存在但未运行；MtStore 无数据则 RLS 联调、经营看板均空转", "演示/验收时无数据可查，功能无法验证", "待处理", "P1 运行时验证阶段立即执行 seed.py"),
    (3, "处方安全", "高", "合理用药引擎是玩具 Mock：仅 4 条规则；prescriptions.create 调用时 patient 传空 {}，禁忌与剂量告警不触发", "处方安全合规不满足，存在医疗风险", "待处理", "P2 接 HttpRxEngine 真实供应商 + patient 信息传入"),
    (4, "前端覆盖", "高", "前端覆盖极窄：admin-web 仅 7 目录；调理师工作台、医院管理、平台监管无前端；mp-taro 仅 5 空壳页", "产品不可对外运营，用户体验无法验证", "待处理", "P3 业务前台补位（小程序+工作台+监管看板）"),
    (5, "业务完整", "高", "互联网医院主线不完整：ih/ 缺科室/药房/药师审核配置；3.7 监管看板缺失；支付/短信第三方对接缺失", "互联网医院主线业务流不闭环", "待处理", "P2/P3 阶段逐项补齐"),
    (6, "安全合规", "高", "等保三级未测评，安全控制未闭环：双因子、落库敏感字段加密（KMS）、审计防篡改（WORM）均未做", "安全测评不通过，无法上线", "待处理", "P4 安全闭环 + P0 等保测评排期"),
    (7, "性能容灾", "中", "性能与容灾未验证：无压测；ClickHouse 仅聚合层占位；AI 推理响应时间未测", "高并发场景可能崩溃，SLA 无法保障", "待处理", "P5 压测 + CH验证 + AI推理测试"),
    (8, "AI真实性", "中", "AI 推理真实性未验证：repurchase/risk 是否真实训练产出、pred_log 是否真实落库未经运行时验证", "AI能力无法确认，影响数据中台核心壁垒", "待处理", "P1 运行时验证 pred_log 落库"),
    (9, "合规资质", "高", "合规硬门槛（组织流程）全部未启动：等保三级测评、医疗机构执业许可证、互联网医院准入、药品经营许可证、ICP证、医师多点执业备案", "阻塞上线，无资质无法合法运营", "待处理", "P0 最高优先级立即启动全部申办"),
]

for i, item in enumerate(risks):
    r = 6 + i
    write_data_row(ws4, r, list(item), alt=(i % 2 == 1))
    # 等级颜色
    lvl_cell = ws4.cell(row=r, column=3)
    lvl_val = item[2]
    if lvl_val == "高":
        lvl_cell.fill = FILL_RED
        lvl_cell.font = Font(name="Arial", size=10, bold=True, color="A93226")
    elif lvl_val == "中":
        lvl_cell.fill = FILL_AMBER
        lvl_cell.font = Font(name="Arial", size=10, bold=True, color="B7780B")
    else:
        lvl_cell.fill = FILL_GREEN
        lvl_cell.font = Font(name="Arial", size=10, bold=True, color="1B7A3D")
    lvl_cell.alignment = A_CENTER
    # 状态颜色
    st_cell = ws4.cell(row=r, column=7)
    st_cell.fill = FILL_AMBER
    st_cell.font = Font(name="Arial", size=10, bold=True, color="B7780B")
    ws4.row_dimensions[r].height = 50

# 风险汇总
risk_sum = 6 + len(risks) + 1
high_count = sum(1 for r in risks if r[2] == "高")
mid_count = sum(1 for r in risks if r[2] == "中")
merge_cell(ws4, f"B{risk_sum}:C{risk_sum}", "风险汇总", F_TD_B, FILL_LIGHT, A_CENTER, B_ALL)
merge_cell(ws4, f"D{risk_sum}:D{risk_sum}", f"高: {high_count}", Font(name="Arial", size=10, bold=True, color="A93226"), FILL_RED, A_CENTER, B_ALL)
merge_cell(ws4, f"E{risk_sum}:E{risk_sum}", f"中: {mid_count}", Font(name="Arial", size=10, bold=True, color="B7780B"), FILL_AMBER, A_CENTER, B_ALL)
merge_cell(ws4, f"F{risk_sum}:H{risk_sum}", f"共 {len(risks)} 项风险 · 其中 {high_count} 项高风险阻塞上线", F_TD_B, FILL_LIGHT, A_CENTER, B_ALL)
ws4.row_dimensions[risk_sum].height = 28

# ─────────────────────────────────────────────────────
# Sheet 5: 资源需求
# ─────────────────────────────────────────────────────
ws5 = wb.create_sheet("资源需求")
ws5.sheet_view.showGridLines = False
set_col_widths(ws5, [3, 16, 14, 12, 35, 14, 14, 3])

for r in range(1, 4):
    for c in range(1, 9):
        ws5.cell(row=r, column=c).fill = FILL_DARK
ws5.row_dimensions[1].height = 8
ws5.row_dimensions[2].height = 36
ws5.row_dimensions[3].height = 22
merge_cell(ws5, "B2:G2", "资源需求与配置计划", F_TITLE, FILL_DARK, A_CENTER)
merge_cell(ws5, "B3:G3", "基于 P0–P5 优先级排期 · 按角色/技能/工时/优先级分解", F_SUB, FILL_DARK, A_CENTER)

write_header_row(ws5, 5, ["角色/资源", "需求类型", "数量/规格", "需求描述", "优先级", "需求时间"])

resources = [
    ("合规专员", "人力", "1-2人", "等保三级定级备案 + 医疗机构执业许可证 + 互联网医院准入 + 药品经营许可证 + ICP证申办", "P0", "立即"),
    ("法律顾问", "外部", "1人（顾问）", "互联网医疗资质合规咨询 + 医师多点执业备案流程指导", "P0", "立即"),
    ("后端工程师", "人力", "2人", "运行时验证（alembic/seed/RLS联调）+ 第三方对接（HttpRxEngine/微信支付/短信）+ 安全闭环", "P1-P4", "第1-4迭代"),
    ("前端工程师", "人力", "3人", "小程序2C/医生端业务流 + 调理师工作台 + 医院管理补齐 + 平台监管看板", "P3", "第3-4迭代"),
    ("AI工程师", "人力", "1人", "AI推理真实性验证 + pred_log落库验证 + 性能测试", "P1/P5", "第1/5迭代"),
    ("测试工程师", "人力", "1人", "压测（Locust/k6）+ ClickHouse验证 + AI推理响应测试 + WAF/CSRF验证", "P5", "第5迭代"),
    ("运维工程师", "人力", "1人", "基础设施部署验证 + nginx配置补齐 + WAF/CSRF部署 + 监控告警", "P3/P5", "贯穿"),
    ("UI/UX设计师", "人力", "1人", "小程序 + 调理师工作台 + 监管看板界面设计", "P3", "第3迭代"),
    ("PostgreSQL", "基础设施", "生产环境", "生产级 PG + pgvector 实例（含备份/高可用）", "P1", "第1迭代"),
    ("ClickHouse", "基础设施", "生产环境", "生产级 ClickHouse 集群（OLAP 经营宽表）", "P1", "第1迭代"),
    ("合理用药API", "第三方服务", "1套", "真实合理用药供应商 API 对接（处方前置校验）", "P2", "第2迭代"),
    ("微信支付", "第三方服务", "1套", "微信支付商户号 + 接口对接", "P2", "第2迭代"),
    ("短信服务", "第三方服务", "1套", "短信验证码/通知服务（阿里云/腾讯云）", "P2", "第2迭代"),
    ("KMS密钥管理", "基础设施", "1套", "密钥管理服务（敏感字段加密存储）", "P4", "第4迭代"),
    ("等保测评机构", "外部", "1家", "等保三级测评服务供应商", "P0", "立即"),
]

for i, item in enumerate(resources):
    r = 6 + i
    write_data_row(ws5, r, list(item), alt=(i % 2 == 1))
    # 优先级颜色
    pri_cell = ws5.cell(row=r, column=5)
    pri_val = item[4]
    if pri_val.startswith("P0"):
        pri_cell.fill = FILL_RED
        pri_cell.font = Font(name="Arial", size=10, bold=True, color="A93226")
    elif "P1" in pri_val:
        pri_cell.fill = FILL_AMBER
        pri_cell.font = Font(name="Arial", size=10, bold=True, color="B7780B")
    elif "P3" in pri_val:
        pri_cell.fill = FILL_LIGHT
        pri_cell.font = Font(name="Arial", size=10, bold=True, color=C_DARK)
    else:
        pri_cell.fill = FILL_GREEN
        pri_cell.font = Font(name="Arial", size=10, bold=True, color="1B7A3D")
    pri_cell.alignment = A_CENTER
    ws5.row_dimensions[r].height = 36

# 汇总
res_sum = 6 + len(resources) + 1
merge_cell(ws5, f"B{res_sum}:G{res_sum}",
           f"共 {len(resources)} 项资源需求 · 其中 {sum(1 for r in resources if r[4].startswith('P0'))} 项需立即启动（P0 阻塞上线）",
           F_TD_B, FILL_LIGHT, A_CENTER, B_ALL)
ws5.row_dimensions[res_sum].height = 28

# ─────────────────────────────────────────────────────
# Sheet 6: 技术架构对照
# ─────────────────────────────────────────────────────
ws6 = wb.create_sheet("技术架构对照")
ws6.sheet_view.showGridLines = False
set_col_widths(ws6, [3, 20, 30, 30, 12, 3])

for r in range(1, 4):
    for c in range(1, 7):
        ws6.cell(row=r, column=c).fill = FILL_DARK
ws6.row_dimensions[1].height = 8
ws6.row_dimensions[2].height = 36
ws6.row_dimensions[3].height = 22
merge_cell(ws6, "B2:E2", "技术架构设计方案 V1.2 对照", F_TITLE, FILL_DARK, A_CENTER)
merge_cell(ws6, "B3:E3", "逐章节验证 · 确认代码实现与技术架构方案的符合度", F_SUB, FILL_DARK, A_CENTER)

write_header_row(ws6, 5, ["架构章节", "要求", "实现情况", "判定"])

arch_items = [
    ("10.2 RBAC 六角色", "JWT + 角色鉴权", "core/deps.require_role、core/security", "✅"),
    ("11.2 plat 资产/模型目录", "模型与资产 CRUD", "P3 真实化", "✅"),
    ("13 数据分级 L1-L4 / 隐私计算", "分级建模、联邦学习", "sensitivity_level 建模；隐私计算仅规划", "⚠️"),
    ("14 第三方对接（支付/短信/合理用药）", "第三方可替换", "合理用药=抽象层+Mock；微信支付/短信未做", "⚠️"),
    ("15.2 RLS 行级隔离", "按门店隔离", "store_scope + 8端点注入 + JWT store_id", "✅"),
    ("15.4 AI 可解释/反馈闭环", "依据 + 采纳回写", "ai-service/rag + pred_logs.adopted", "✅"),
    ("15.9 等保三级", "测评 + 资质", "清单已出；测评/资质未启动", "⚠️"),
]
for i, item in enumerate(arch_items):
    r = 6 + i
    write_data_row(ws6, r, list(item), alt=(i % 2 == 1), status_col=3)
    ws6.row_dimensions[r].height = 32

# ─────────────────────────────────────────────────────
# 保存
# ─────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"OK: {OUTPUT}")
